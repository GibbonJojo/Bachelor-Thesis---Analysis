import openpyxl
from openpyxl.styles import PatternFill
from Get_Data_OneRow import get_data
import numpy as np

"""
You don't have to change anything in this file, except the following path.
Specify the path in which the Excel Workbook should be saved into.
Please mind, that there has to be a "r" in front of the quotes!
"""
WB_PATH = r"C:\Users\jojo\OneDrive\Dokumente\Uni\BACHELORARBEIT\Analysis"

"""
Before you run the script now, open "Get_Data_OneRow.py" with a text editor (IDLE) and follow the instructions there.
"""

# Create the Workbook
workbook = openpyxl.Workbook()
posner_sheet = workbook.active
posner_sheet.title = "Posner"
posner_online_sheet = workbook.create_sheet("Posner_Online")
director_sheet = workbook.create_sheet("Director")
director_online_sheet = workbook.create_sheet("Director_Online")

# fill the sheet
def fill_workbook(posner_data, director_data, posner_online, director_online):
    # write posner_data into posner_sheet
    for i, key in enumerate(posner_data, 1):  # fills the first row with the respective names
        posner_sheet.cell(row=1, column=i, value=key)

    for i, column in enumerate(posner_data.values(), 1):  # fills the data in.
        for x, num in enumerate(column, 2):
            posner_sheet.cell(row=x, column=i, value=num)

        # Fill in means and stds
        if i == 1:
            continue
        else:
            posner_sheet.cell(row=len(column)+3, column=i, value=np.average(column))
            posner_sheet.cell(row=len(column)+4, column=i, value=np.std(column))

###########################
    # Write director_data into director_sheet
    for i, key in enumerate(director_data, 1):
        director_sheet.cell(row=1, column=i, value=key)

    for i, column in enumerate(director_data.values(), 1):
        for x, num in enumerate(column, 2):
            director_sheet.cell(row=x, column=i, value=num)

        if i == 1:
            continue
        else:
            director_sheet.cell(row=len(column)+3, column=i, value=np.average(column))
            director_sheet.cell(row=len(column)+4, column=i, value=np.std(column))
###########################
    # Write posner_online_data into posner_online_sheet
    for i, key in enumerate(posner_online, 1):
        posner_online_sheet.cell(row=1, column=i, value=key)

    for i, column in enumerate(posner_online.values(), 1):
        for x, num in enumerate(column, 2):
            posner_online_sheet.cell(row=x, column=i, value=num)
        if i == 1:
            continue
        else:
            posner_online_sheet.cell(row=len(column)+3, column=i, value=np.average(column))
            posner_online_sheet.cell(row=len(column)+4, column=i, value=np.std(column))
###########################
    # Write director_online_data into director_online_sheet
    for i, key in enumerate(director_online, 1):
        director_online_sheet.cell(row=1, column=i, value=key)

    for i, column in enumerate(director_online.values(), 1):
        for x, num in enumerate(column, 2):
            director_online_sheet.cell(row=x, column=i, value=num)
        if i == 1:
            continue
        else:
            director_online_sheet.cell(row=len(column)+3, column=i, value=np.average(column))
            director_online_sheet.cell(row=len(column)+4, column=i, value=np.std(column))

def main():
    posner_data = get_data("posner")  # extract the data from posner test
    director_data = get_data("director")  # extract the data from director test
    posner_data_online = get_data("posner_online")  # extract the online posner data
    director_data_online = get_data("director_online")  # extract the online director data

    # if you want to inspect, what the data looks like, uncomment the next lines
    # if you want to inspect the other datasets, replace posner_data with the name of the data, you want to inspect
    # for data in posner_data:
    #     print(data)
    #     print(posner_data.get(data))
    #     print()

    fill_workbook(posner_data, director_data, posner_data_online, director_data_online)  # fill the excel workbook with the data

    workbook.save(WB_PATH + "\\Behavioral.xlsx")  # save the workbook


if __name__ == "__main__":
    main()
