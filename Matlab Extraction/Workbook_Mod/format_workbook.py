from openpyxl.styles import PatternFill


# Change a value in the excel sheet to string.
def as_text(value):
    return str(value) if value is not None else ""


def format_sheet(sheet):
    """Format the excel sheet for better readability"""

    # Fit the Column width to the input
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            length = len(str(cell.value) if cell.value is not None else "Mean_Down ")
    #        sheet.column_dimensions[cell.column].width = length
    #sheet.column_dimensions["A"].width = len("Down_mean   ")

    # Freeze the first row and column
    sheet.freeze_panes = "B2"

    # Define the color grey for later filling of every other line
    grey = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    x = 0
    # Change the number format to less decimal places
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, float) or isinstance(cell.value, int):
                cell.number_format = "0.000"

                # Color every other row in light grey
        if x == 0:
            for cell in row:
                cell.fill = grey
                x = 1
        elif x == 1:
            x = 0
            continue
