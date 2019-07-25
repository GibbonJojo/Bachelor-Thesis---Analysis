import openpyxl


def init_workbook():
    # Initialize the Excel Workbook
    # Create the Workbook for the excel sheet
    workbook = openpyxl.Workbook()

    # Create the excel sheets
    behavioral_sheet = workbook.active
    behavioral_sheet.title = "Behavioral"

    workbook.create_sheet("Time_above_thresh_trials")
    workbook.create_sheet("Time_above_thresh_run")
    workbook.create_sheet("Time_above_thresh_sess")

    workbook.create_sheet("Time_above_init_thresh_trials")
    workbook.create_sheet("Time_above_init_thresh_run")
    workbook.create_sheet("Time_above_init_thresh_sess")

    workbook.create_sheet("Performance_mean_trials")
    workbook.create_sheet("Performance_mean_run")
    workbook.create_sheet("Performance_mean_sess")

    workbook.create_sheet("Performance_median_trials")
    workbook.create_sheet("Performance_median_run")
    workbook.create_sheet("Performance_median_sess")

    workbook.create_sheet("Performance_std_trials")
    workbook.create_sheet("Performance_std_run")
    workbook.create_sheet("Performance_std_sess")

    workbook.create_sheet("Baseline_median_trials")
    workbook.create_sheet("Baseline_median_run")
    workbook.create_sheet("Baseline_median_sess")

    workbook.create_sheet("Baseline_std_trials")
    workbook.create_sheet("Baseline_std_run")
    workbook.create_sheet("Baseline_std_sess")

    return workbook
